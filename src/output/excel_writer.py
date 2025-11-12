"""
ExcelWriter - Excelファイルへの書き込み
"""

import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../..'))

from src.data_structures import TruthTableData, IOTableData


def ensure_directory(dirpath):
    """ディレクトリが存在しない場合は作成"""
    if dirpath and not os.path.exists(dirpath):
        os.makedirs(dirpath)


class ExcelWriter:
    """Excelファイルへの書き込みクラス"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def write_truth_table(self, data: TruthTableData, filepath: str) -> None:
        """真偽表をExcelに書き込み"""
        self.logger.info(f"真偽表をExcelに書き込み: {filepath}")
        
        ensure_directory(os.path.dirname(filepath))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "真偽表"
        
        # ヘッダー
        headers = ['No.', '真偽', '判定文', '期待値']
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # データ
        for row_idx, tc in enumerate(data.test_cases, 2):
            ws.cell(row=row_idx, column=1, value=tc.no)
            ws.cell(row=row_idx, column=2, value=tc.truth)
            ws.cell(row=row_idx, column=3, value=tc.condition)
            ws.cell(row=row_idx, column=4, value=tc.expected)
        
        # 列幅調整
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        
        wb.save(filepath)
        self.logger.info(f"真偽表の書き込みが完了: {len(data.test_cases)}行")
    
    def write_io_table(self, data: IOTableData, filepath: str) -> None:
        """I/O表をExcelに書き込み"""
        self.logger.info(f"I/O表をExcelに書き込み: {filepath}")
        
        ensure_directory(os.path.dirname(filepath))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "IO表"
        
        excel_data = data.to_excel_format()
        
        if len(excel_data) < 2:
            self.logger.warning("I/O表のデータが不足しています")
            return
        
        # スタイル定義
        header_font = Font(bold=True, size=11)
        input_fill = PatternFill(start_color="E7F4FF", end_color="E7F4FF", fill_type="solid")
        output_fill = PatternFill(start_color="FFE7E7", end_color="FFE7E7", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # データを書き込み
        for row_idx, row_data in enumerate(excel_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # ヘッダー行1（input/output）
                if row_idx == 1:
                    cell.font = header_font
                    if value == 'input':
                        cell.fill = input_fill
                    elif value == 'output':
                        cell.fill = output_fill
                
                # ヘッダー行2（変数名）
                elif row_idx == 2:
                    cell.font = header_font
                    
                    # input列かoutput列かを判定
                    if col_idx > 2:
                        header1_value = excel_data[0][col_idx - 1]
                        if header1_value == 'input':
                            cell.fill = input_fill
                        elif header1_value == 'output':
                            cell.fill = output_fill
        
        # 列幅を調整
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        
        for col_idx in range(3, len(excel_data[0]) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 12
        
        # 1行目と2行目をマージ（No、テスト名列）
        if len(excel_data) >= 2:
            ws.merge_cells('A1:A2')
            ws.merge_cells('B1:B2')
        
        wb.save(filepath)
        self.logger.info(f"I/O表の書き込みが完了: {len(data.test_data)}行")
